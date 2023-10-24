import tkinter as tk
import pandas as pd
import webbrowser
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from compare_neotech import compare_neotech
from perform_vlookup import perform_vlookup


def format_headers_in_excel(filename):
    # Load the workbook
    workbook = load_workbook(filename)

    headers_to_color = ['PSOFT PART', 'PSID CT', 'QUOTED MFG', 'QUOTED PART', 'PART CLASS', 'AML CPN_MFGID', 'NAME',
                        'AML CPN_MFGNUM',
                        'AML CPN_MFGPARTNUM', 'LIFECYCLE STATUS']
    color_list = ["fffa9e", "fffa9e", "fffa9e", "fffa9e", "fffa9e", "f3f800", "f3f800", "f3f800", "f3f800", "f3f800"]
    header_color_mapping = dict(zip(headers_to_color, color_list))

    general_headers_to_color = ['COMPANY', 'PLANT', 'SITE NAME', 'PARTNUM', 'BUYERID', 'IUM', 'PUM', 'NCNR FLAG',
                                'COMM CODE', 'CUSTOMER PN', 'REFERENCE', 'VENDORNUM', 'VENDORID', 'VENDORNAME',
                                'CONSOLIDATED NAME', 'CONSOLIDATED NAME 2', 'GROUPDESC', 'MFGNUM', 'MFGID',
                                'MFGPARTNUM', 'BASEUNITPRICE', 'EFFECTIVEDATE', 'EXPIRATIONDATE', 'CLASSID',
                                'PURPOINT', 'PARTDESCRIPTION', 'PRODCODE', 'ACTIVE Y OR N', 'CONSOLIDATED CUST NAME',
                                'MRPSHARE_C', 'ASOF', 'MINORDERQTY', 'MFGLOTMULTIPLE', 'MINMFGLOTSIZE', 'LEADTIME',
                                'PRICELIST_LT', '90 DAY DEMAND', 'TOTAL DEMAND', 'SUMOFONHANDQTY']

    general_color = "00abee"  # Light blue

    align_wrap = Alignment(wrap_text=True)

    # Loop through all sheets in the workbook
    for sheet in workbook.worksheets:
        for cell in sheet[1]:  # Headers are in the first row
            cell.alignment = align_wrap
            # Color the specified headers in the "Full File Without Dupes" sheet
            if sheet.title == 'Full File Without Dupes':
                if cell.value in header_color_mapping:
                    cell.fill = PatternFill(start_color=header_color_mapping[cell.value],
                                            end_color=header_color_mapping[cell.value],
                                            fill_type="solid")
                elif cell.value in general_headers_to_color:
                    cell.fill = PatternFill(start_color=general_color, end_color=general_color, fill_type="solid")

    workbook.save(filename)


def save_to_excel(original_data, unique_data, removed_from_prev_data, output_name):
    with pd.ExcelWriter(output_name, engine='openpyxl') as writer:
        original_data.to_excel(writer, sheet_name='Full Original File', index=False)
        unique_data.to_excel(writer, sheet_name='Full File Without Dupes', index=False)
        removed_from_prev_data.to_excel(writer, sheet_name='Removed from Prev File', index=False)


# Create the GUI window
window = tk.Tk()

# Set the window geometry to a larger size
window.geometry("1000x500")


def open_readme_link():
    webbrowser.open('https://github.com/nabilcanan/Partnership_Sort_NeoTech/blob/main/README.md',
                    new=2)  # new=2 ensures the link opens in a new window.


# Add a title label
title_label = tk.Label(window, text="Comparing NeoTech Contract Files",
                       font=("Microsoft YaHei", 28, "bold", "underline"), foreground="red")
title_label.grid(row=0, column=0, pady=20)

# Add instructions label
instructions_label = tk.Label(window,
                              text="Instructions:\n"
                                   "1. Select your most recent NeoTech Contract File.\n"
                                   "2. Select the previous Neotech Contract File.\n"
                                   "3. Finally choose where you'd like to save your final workbook\n",
                              font=("Microsoft YaHei", 18))
instructions_label.grid(row=1, column=0, pady=10)

# Create a frame for the first button
button_frame1 = tk.Frame(window)
button_frame1.grid(row=2, column=0, pady=10)

# Add a button to trigger file selection and comparison
compare_button = tk.Button(button_frame1, text='Compare NeoTech Files', command=compare_neotech,
                           font=("Microsoft YaHei", 20, "bold"), bg="red", fg="white")
compare_button.pack(fill='both')

vlookup_button = tk.Button(button_frame1, text='Perform V-lookup', command=perform_vlookup,
                           font=("Microsoft YaHei", 20, "bold"), bg="blue", fg="white")
vlookup_button.pack(fill='both')

# Create a frame for the README link button
button_frame2 = tk.Frame(window)
button_frame2.grid(row=3, column=0, pady=10)

readme_button = tk.Button(button_frame1, text='Open README', command=open_readme_link,
                          font=("Microsoft YaHei", 20, "bold"), bg="blue", fg="white")
readme_button.pack(fill='both')

# Configure the grid to expand
for i in range(6):
    window.grid_rowconfigure(i, weight=1)
window.grid_columnconfigure(0, weight=1)

# Run the GUI window
window.mainloop()
