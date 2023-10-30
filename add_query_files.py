import pandas as pd
import openpyxl
from tkinter import filedialog, messagebox
from openpyxl.styles import PatternFill


def add_excel_files_to_workbook():
    # Ask the user to select the target workbook
    target_workbook_path = filedialog.askopenfilename(title="Select the target workbook",
                                                      filetypes=[("Excel files", ("*.xlsx", "*.xls"))])
    if not target_workbook_path:
        return "No target workbook selected."

    # If the file is an .xls, convert it to .xlsx
    if target_workbook_path.endswith('.xls'):
        # Read the xls file with pandas
        xls_data = pd.read_excel(target_workbook_path, engine='xlrd')

        # Convert .xls to .xlsx by saving with a new extension
        target_workbook_path = target_workbook_path.replace('.xls', '_converted.xlsx')
        xls_data.to_excel(target_workbook_path, engine='openpyxl', index=False)

    main_wb = openpyxl.load_workbook(target_workbook_path)

    filepaths = filedialog.askopenfilenames(title="Select the Excel files to add",
                                            filetypes=[("Excel files", "*.xlsx")])
    if not filepaths:
        return "No Excel files selected."

    colored_columns = ['PSoft Part', 'PSID CT', 'Quoted Mfg', 'Quoted Part', 'Part Class']
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue

    for filepath in filepaths:
        if 'award' in filepath.lower():
            sheet_name = 'Awards'
        elif 'snd' in filepath.lower():
            sheet_name = 'SND'
        elif 'vpc' in filepath.lower():
            sheet_name = 'VPC'
        elif 'sales' in filepath.lower():
            sheet_name = 'Sales'
        elif 'backlog' in filepath.lower():
            sheet_name = 'Backlog'
        else:
            continue  # Skip files that don't meet any criteria

        data_wb = openpyxl.load_workbook(filepath)
        data_ws = data_wb.active
        new_ws = main_wb.create_sheet(sheet_name)

        for idx, row in enumerate(data_ws):
            for cell in row:
                new_cell = new_ws[cell.coordinate]
                new_cell.value = cell.value

                # Apply formatting to header row
                if idx == 0:
                    new_cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                    if cell.value in colored_columns:
                        new_cell.fill = header_fill

        # Freeze panes just below the first row and at column J
        new_ws.freeze_panes = new_ws["K2"]

    save_path = filedialog.asksaveasfilename(title="Save Workbook As",
                                             filetypes=[("Excel files", ("*.xlsx", "*.xls"))],
                                             defaultextension=".xlsx")

    if save_path:
        main_wb.save(save_path)
        messagebox.showinfo("Success", f"Data saved to {save_path}.")
        return f"Data saved to {save_path}."
    else:
        return "Save operation canceled."

# Test
# print(add_excel_files_to_workbook())
